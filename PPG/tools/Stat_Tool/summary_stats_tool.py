# https://realpython.com/openpyxl-excel-spreadsheets-python
import pysftp
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook,load_workbook
from collections import Counter
import json
import os

RAW_DATA_PATH = "24EI/RAW_DATA"

host_name = "datastore.oucru.org"
user_name = "username@oucru.org"
password = "password"

cnopts = pysftp.CnOpts()
cnopts.hostkeys = None      # May cause Man-in-the-Middle attack

file_names = []
dir_names = []
un_name = []
directory_structure = []

def write_log(file_idx,patient_detail_number):
    """
    Write the log file to resume the running
    :param file_idx: The current reading file
    :param patient_detail_number:  The current row of the patient infomration in 'patient detail' sheet
    :return:
    """
    data = {}
    data['file_number'] = file_idx
    data['patient_detail_number'] = patient_detail_number
    with open('log.txt', 'w') as outfile:
        json.dump(data, outfile)

def store_files_name(fname):
    file_names.append(fname)

def store_dir_name(dirname):
    dir_names.append(dirname)

def store_other_file_types(name):
    un_name.append(name)

def do_nothing(inp):
    pass

def summary_sheet(ws):
    """
    Fill in detail for summary sheet. Refer to the template
    :param ws: the summary sheet of the workbook
    :return:
    """

    ppg_wearable_files = [f for f in file_names if (('+' in f) or ('PPG' in f)) and 'csv' in f]
    ppg_days = set([day.split("/")[-2] for day in ppg_wearable_files])

    ws["B3"] = "Number of Patients"
    ws["C3"] = 110

    ws["B4"] = "Number of recorded days"
    ws["C4"] = len(ppg_days)

    ws["B5"] = "Number of missing days (PPG)"
    ws["C5"] = 110*2 - len(ppg_days)

    ws["B6"] = "Number of missing ECG"
    ws["C6"] = "TO DO"

    ws["B7"] = "Number of day having splited PPG record"
    ws["C7"] = len(ppg_wearable_files)- len(ppg_days)

    ws["B8"] = "Day having splited PPG record"
    ws["C8"] = ""

    return ws

def list_file_by_day(ws,day_list,row_index,patient_detail_row_index,day_number=0):
    """
    Input all recorded file in the given day to the work sheet
    :param ws: The patient list sheet
    :param day_list: list of days having the recorded data
    :param row_index: The current row of the cusor in the patient summary sheet
    :param patient_detail_row_index: The row of the hyperlink to the patient detail sheet
    :param day_number: Day of record - Day 1, Day 5 or other Day
    :return:
    row_index
    patient_detail_row_index
    """
    if len(day_list) < 1 and day_number!=0:
        if day_number !=1:
            day_number = 5
        ws["D" + str(row_index)] = "DAY "+str(day_number)
        ws["F" + str(row_index)] = "Missing File"
        row_index = row_index + 1
        # return row_index, patient_detail_row_index
    else:
        for day_file in day_list:
            ws["D" + str(row_index)] = day_file.split("/")[3]
            cell = ws.cell(row_index, 6)
            cell.value = '=HYPERLINK("#\'Patient Detail\'!A' + str(patient_detail_row_index) + \
                         '","' + day_file.split("/")[-1] + '")'

            patient_detail_row_index = patient_detail_row_index + 18
            row_index = row_index + 1
    return row_index, patient_detail_row_index

def patient_list_sheet(ws):
    """
    Fill in 'patient list' sheet information. Refer to the template.
    The structure start with the patient ID  - Days recorded - Filename in each day (or missing)
    :param ws: The 'patient list' sheet
    :return:
    """
    day_record_folders = [d for d in dir_names if ("DAY" in d.upper())]
    # GET TOTAL PATIENT BY CHECK THE FOLDER FORMAT "24EI - XXX"
    patient_folder_list = [p.split("/")[2][:10] for p in day_record_folders]
    patient_counter = Counter(patient_folder_list)
    ws["B4"] = "Patient ID"
    ws["D4"] = "Day"
    ws["F4"] = "Link to PPG stats"
    row_index = 5
    patient_detail_row_index = 1

    for patient in patient_counter.keys():
        ws["B"+str(row_index)] = patient
        # if patient in day_record_folders:
        day_1_list = [p_file for p_file in ppg_wearable_files
                      if (patient in p_file) and (("DAY1" in p_file.upper()) or ("DAY 1") in p_file.upper())]
        day_5_list = [p_file for p_file in ppg_wearable_files
                      if (patient in p_file) and (("DAY5" in p_file.upper()) or ("DAY 5") in p_file.upper())]
        other_day_list = [p_file for p_file in ppg_wearable_files
                      if (patient in p_file) and (("DAY1" not in p_file.upper()) and ("DAY 1") not in p_file.upper())
                                             and (("DAY5" not in p_file.upper()) and ("DAY 5") not in p_file.upper())
                                             and ("DAY" in p_file.upper())]

        row_index, patient_detail_row_index = list_file_by_day(ws,day_1_list,row_index,patient_detail_row_index,1)
        row_index, patient_detail_row_index = list_file_by_day(ws,day_5_list, row_index, patient_detail_row_index,5)
        row_index, patient_detail_row_index = list_file_by_day(ws,other_day_list, row_index, patient_detail_row_index)

    return ws

def write_describe(fname, df_describe,ws,patient_number):
    """
    Input data statistical information
    :param fname: the csv data file
    :param df_describe: Dataframe contain the statisical information given by pandas describe
    :param ws: The patient detail sheet
    :param patient_number: Patient number
    :return:
    """
    ws["A"+str(patient_number+1)] = "File name"
    ws["B" + str(patient_number + 1)] = fname.split("/")[-1]

    ws["A"+str(patient_number+2)] = "Path"
    ws["B" + str(patient_number + 2)] = "/".join(fname.split("/")[:-1])

    ws["A"+str(patient_number+4)] = "Patient"
    ws["B" + str(patient_number + 4)] = fname.split("/")[2]

    ws["C"+str(patient_number+4)] = "Day"
    ws["D" + str(patient_number + 4)] = fname.split("/")[3]

    #================================================================
    ite = 6
    for stat in df_describe.index:
        ws["C" + str(patient_number + ite)] = stat
        ws["D" + str(patient_number + ite)] = df_describe["PLETH"][stat]
        ite = ite+1

    ws["C" + str(patient_number + 14)] = "Invalid SpO2 ratio"
    ws["D" + str(patient_number + 8)] = ""

    ws["C" + str(patient_number + 15)] = "Invalid Pulse ratio"
    ws["D" + str(patient_number + 8)] = ""

    ws["C" + str(patient_number + 16)] = "Poor Pleth ratio"
    ws["D" + str(patient_number + 8)] = ""
    return ws

def fill_patient_detail(workbook, ppg_wearable_files,file_number =0,patient_number=0):
    """
    Fill in the patient detail sheet. Refer to the template
    :param workbook:
    :param ppg_wearable_files:
    :param file_number:
    :param patient_number:
    :return:
    """
    for fname in tqdm(ppg_wearable_files[file_number:]):
        sftp = pysftp.Connection(host=host_name, username=user_name, password=password,
                       cnopts=cnopts)
        sftp.timeout = 10000
        # f = sftp.open(fname)
        # df = pd.read_csv(f,error_bad_lines=False)
        sftp.get(fname, 'temp_read.csv')
        df = pd.read_csv("temp_read.csv", error_bad_lines=False)
        df = df.apply(pd.to_numeric, errors='coerce').dropna()
        summary = df.describe()

        sheet = workbook.active
        write_describe(fname, summary, sheet, patient_number)
        patient_number = patient_number + 18
        file_number = file_number+1
        write_log(file_number,patient_number)
        workbook.save(filename="Signal_Summary_Stats.xlsx")
        sftp.close()

"""
=======================================================================================================
"""
if __name__ == "__main__":
    print("Connection succesfully stablished ... ")
    sftp = pysftp.Connection(host=host_name, username=user_name, password=password,
                                             cnopts=cnopts)
    sftp.timeout = 10000
    # Obtain structure of the remote directory '/var/www/vhosts'
    directory_structure = sftp.listdir_attr("24EI/RAW_DATA")
    sftp.walktree(RAW_DATA_PATH,fcallback=store_files_name,dcallback=store_dir_name,ucallback=do_nothing,
                                          recurse=True) #files.append
    print(file_names)
    ppg_wearable_files = [f for f in file_names if (('+' in f) or ('PPG' in f)) and 'csv' in f]

    print(dir_names)
    if os.path.exists("Signal_Summary_Stats.xlsx"):
        workbook = load_workbook("Signal_Summary_Stats.xlsx")
    else:
        workbook = Workbook()
        workbook.active = 0
        ws = workbook.active
        ws.title = "Summary"
        workbook.create_sheet("Patient list")
        workbook.create_sheet("Patient Detail")

    # Prepare data for summary Sheet
    workbook.active = 0
    ws = summary_sheet(workbook.active)

    workbook.active = 1
    patient_list_sheet(workbook.active)

    workbook.active = 2

    if not os.path.exists("log.txt"):
        write_log(0, 0)

    read_file = open("log.txt", "r")
    log_content = json.load(read_file)
    file_number = log_content["file_number"]
    current_patient_idx = log_content["patient_detail_number"]

    fill_patient_detail(workbook, ppg_wearable_files, file_number, patient_number=current_patient_idx)

    workbook.save(filename="Signal_Summary_Stats.xlsx")